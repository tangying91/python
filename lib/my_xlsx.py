import datetime
import warnings

# 过滤无意义的警告
from openpyxl import load_workbook

warnings.filterwarnings('ignore')


# 根据header找到列的位置
# header默认在第一行
def get_header_column_idx(book_sheet, header):
    for row in book_sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if header == cell.value:
                return cell.column
    return -1


# 获取某一列的不重复数据
def get_no_repeat_column_data(book_sheet, title):
    column = get_header_column_idx(book_sheet, title)
    datas = []
    for row in book_sheet.iter_rows(min_row=2):
        datas.append(row[column].value)
    return datas


# 获取指定分页，指定列，并过滤关键字的数据
# headers是数组，filters是字典
def filter_sheet_data(book_sheet, headers, filters):
    title_columns = {}
    for header in headers:
        column = get_header_column_idx(book_sheet, header)
        title_columns[column] = header

    datas = []
    for row in book_sheet.iter_rows(min_row=2):
        data = {}
        append_data = True
        all_none = True
        for column, header in title_columns.items():
            value = row[column].value
            data[header] = value

            # 检查是否需要过滤
            if header in filters.keys() and value == filters[header]:
                append_data = False

            # 检查是否整行都是空数据
            if value is not None:
                all_none = False

        if append_data is True and all_none is False:
            datas.append(data)
    return datas


# 设置分页指定列列宽
def set_sheet_column_width(book_sheet, columns, width):
    for column in columns:
        book_sheet.column_dimensions[column].width = width


# 用于Excel表格拆分
# 第一行默认为标题栏，不进行任何处理
# 根据关键字拆分Excel，同时输出新的的拆分后的excel，保留所有格式
# input_file_name 源文件
# split_keys 拆分关键字数组
# keep_sheets 需要保留的分页
# output_file_suffix 输出的新文件后缀
def excel_split(input_file_name, split_keys, keep_sheets, output_file_path, output_file_suffix):
    print(input_file_name, "excel split start..")
    # 记录开始时间
    start_time = datetime.datetime.now()
    for keyword in split_keys:
        # 加载Excel文件，仅数据处理很重要，这个可以规避excel公式问题
        wb = load_workbook(input_file_name, data_only=True)

        # 遍历每个工作表
        for sheet_name in wb.sheetnames:
            # 获取当前工作表
            sheet = wb[sheet_name]

            # 删除不要的分页
            if sheet_name not in keep_sheets:
                wb.remove(sheet)
                continue

            # 定义要删除的行的索引
            rows_to_delete = []

            # 遍历每一行，第一行默认为标题，不处理
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                delete = True
                for cell in row:
                    if str(cell) == keyword:
                        delete = False
                        break
                if delete:
                    rows_to_delete.append(row_index)

            print(keyword, sheet_name, "分页处理完成")
            # 删除要删除的行（注意要从后向前删除）
            for row_index in reversed(rows_to_delete):
                sheet.delete_rows(row_index)

        # 关闭Excel文件
        wb.save(output_file_path + keyword + output_file_suffix + '.xlsx')
        wb.close()
        print(keyword, '全部处理完毕，耗时', (datetime.datetime.now() - start_time).total_seconds(), "秒")


# Excel 对指定列重新编号
# input_file_name 源文件
# header 指定列标题
def excel_renumber(input_file_name, header):
    print(input_file_name, "excel renumber start..")
    # 记录开始时间
    start_time = datetime.datetime.now()
    # 打开表格
    wb = load_workbook(input_file_name)

    # 获取所有表名称
    for sheet_name in wb.sheetnames:
        # 获取当前工作表
        sheet = wb[sheet_name]
        # 获取最大行数
        max_row = sheet.max_row

        # 遍历每一行，为第一列赋予自动编号
        column = get_header_column_idx(sheet, header)
        if column != -1:
            # 因为第一行是标题，所以实际值行号往后加1
            for idx in range(1, max_row):
                sheet.cell(row=idx + 1, column=column).value = idx

    # 保存文件
    wb.save(input_file_name)
    print(input_file_name, '重新编号完成，耗时', (datetime.datetime.now() - start_time).total_seconds(), "秒")

