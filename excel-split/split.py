import os

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter

# 功能介绍
# 根据大区拆分表格，保留原数据格式和内容

# 需要处理的源文件
input_file_name = 'input/2023年Q2业绩综合分析-7.3.xlsx'
# 拆分后的文件名称后缀
output_file_suffix = '-2023年Q2销售业绩结算表-0703'
# 拆分表格的关键字
split_keys = ['苏皖大区', '上海分公司', '北京分公司', '北区', '西区', '物流交通', '南区', '浙江大区', '福建办事处']
# 最终需要保留的数据分页
keep_sheets = ['大区', '办事处', '组长', '个人', 'Q2业绩清单', '剔除清单']

print("\n下面开始拆分表格……")

# 读取表格，处理数据
sheets_dict = pd.read_excel(input_file_name, sheet_name=None)

# 首先拆分数据
for keyword in split_keys:
    # 创建一个空的字典来保存拆分后的工作表数据
    split_sheets_dict = {}

    # 遍历每个工作表
    for sheet_name, df in sheets_dict.items():
        # 筛选含有指定字符串的行数据
        filtered_df = df[df.astype(str).apply(lambda row: any(keyword == cell for cell in row), axis=1)]
        # 将筛选后的数据添加到结果字典中
        split_sheets_dict[sheet_name] = filtered_df

    # 导出拆分后的数据到新的Excel，作为临时数据存储
    with pd.ExcelWriter('tmp/' + f'{keyword}.xlsx', engine='openpyxl') as writer:
        for sheet_name, split_df in split_sheets_dict.items():
            split_df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f'{keyword}' + '数据拆分结束')

print("====================================================")
print("下面开始删除空的分页……")

# 其次处理删除空的分页
for keyword in split_keys:
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook('tmp/' + f'{keyword}.xlsx')

    # 遍历每个工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 判断第二行是否为空或没有数据，或者不在所需的分页里
        row_values = [cell.value for cell in sheet[2]]
        if all(value is None for value in row_values) or sheet_name not in keep_sheets:
            # 删除该工作表
            workbook.remove(sheet)

    # 保存修改后的文件
    workbook.save('tmp/' + f'{keyword}.xlsx')
    print(f'{keyword}' + '空分页处理结束')

print("====================================================")
print("下面开始处理序号……")

# 重新所有的序号标题，生成新文件，并删除临时文件
for keyword in split_keys:
    # 打开 Excel 文件
    input_file = pd.ExcelFile('tmp/' + f'{keyword}.xlsx')

    # # 创建新的 Excel 文件
    output_file = pd.ExcelWriter('output/' + f'{keyword}{output_file_suffix}.xlsx', engine='openpyxl')

    # 遍历每一页
    for sheet_name in input_file.sheet_names:
        # 读取当前页的数据
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # 删除已有的序号列
        if '序号' in df.columns:
            df.drop('序号', axis=1, inplace=True)

        # 添加新的序号列
        df.insert(0, '序号', range(1, len(df) + 1))

        # 写入当前页的数据到新的 Excel 文件中
        df.to_excel(output_file, sheet_name=sheet_name, index=False)

    # 保存并关闭 Excel 文件
    output_file.close()
    input_file.close()

    # # 删除旧文件（如果存在）
    if os.path.exists(input_file):
        os.remove(input_file)
    print(f'{keyword}' + '序号处理结束')

print("====================================================")
print("下面开始渲染格式……")

# 原始的模板Excel文件
template_wb = openpyxl.load_workbook(input_file_name)
print("目标原始文件打开成功")

# 重新所有的序号标题，生成新文件，并删除临时文件
for keyword in split_keys:
    # 打开目标Excel文件
    destination_file = 'output/' + f'{keyword}{output_file_suffix}.xlsx'

    # 加载工作簿
    destination_wb = openpyxl.load_workbook(destination_file)
    print(f'{keyword}' + " 开始渲染")

    # 获取源工作表和目标工作表
    template_sheets = template_wb.sheetnames
    destination_sheets = destination_wb.sheetnames

    # 遍历每个工作表
    for i, template_sheet_name in enumerate(destination_sheets):
        template_sheet = template_wb[template_sheet_name]
        destination_sheet = destination_wb[template_sheet_name]

        # 获取源工作表的列数
        column_count = min(template_sheet.max_column, 50)

        # 复制每一列的样式和列宽到目标工作表
        for col in range(1, column_count + 1):
            col_letter = get_column_letter(col)  # 列索引转换为字母表示的列名，如1->A, 2->B
            source_column = template_sheet[col_letter]
            destination_column = destination_sheet[col_letter]
            for source_cell, destination_cell in zip(source_column, destination_column):
                destination_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    underline=source_cell.font.underline,
                    color=source_cell.font.color
                )

                destination_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    fgColor=source_cell.fill.fgColor,
                    bgColor=source_cell.fill.bgColor
                )

                destination_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )

                destination_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrapText=source_cell.alignment.wrapText
                )

                destination_cell.number_format = source_cell.number_format
            destination_sheet.column_dimensions[col_letter].width = template_sheet.column_dimensions[col_letter].width

    # 保存目标工作簿文件
    destination_wb.save(destination_file)
    print(f'{keyword}' + '渲染格式结束')

print("====================================================")
print("所有数据拆分渲染结束!!!")
