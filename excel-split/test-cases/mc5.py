
from openpyxl import load_workbook, Workbook

def filter_excel(input_file, output_file, keyword):
    # 加载Excel文件
    wb = load_workbook(input_file)

    # 遍历每个工作表
    for ws_name in wb.sheetnames:
        # 获取当前工作表
        ws = wb[ws_name]

        # 定义要删除的行的索引
        rows_to_delete = []

        # 从最后一行开始遍历
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # 检查每个单元格的值是否包含关键字
            if any(keyword != str(cell_value) for cell_value in row):
                # 将要删除的行的索引保存到列表中
                rows_to_delete.append(row_index)

        # 删除要删除的行（注意要从后向前删除）
        for row_index in reversed(rows_to_delete):
            ws.delete_rows(row_index)

    # 保存新的Excel文件
    wb.save(output_file)


# 指定输入文件路径、输出文件路径和关键字
input_file = '../output/西区-2023年Q2销售业绩结算表-0703.xlsx'
output_file = '../output/mc5.xlsx'
keyword = '西区办事处'

# 调用函数进行筛选和保存
filter_excel(input_file, output_file, keyword)
