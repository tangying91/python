import pandas as pd
from openpyxl import load_workbook

# 读取 Excel 文件中的所有工作表
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter

sheets_dict = pd.read_excel('2023年Q1业绩综合分析V4终版-4.11.xlsx', sheet_name=None)

# 定义字符串数组
strings = ['苏皖大区', '上海分公司', '北京分公司', '北区', '西区', '物流交通大区', '南区', '浙江大区']

# # 遍历每个字符串
# for string in strings:
#     # 创建一个空的字典来保存拆分后的工作表数据
#     split_sheets_dict = {}
#
#     # 遍历每个工作表
#     for sheet_name, df in sheets_dict.items():
#         # 筛选含有指定字符串的行数据
#         filtered_df = df[df.astype(str).apply(lambda row: any(string in cell for cell in row), axis=1)]
#         # 将筛选后的数据添加到结果字典中
#         split_sheets_dict[sheet_name] = filtered_df
#
#     # 导出拆分后的数据到新的 Excel 文件，每个工作表分别保存为一个 sheet
#     with pd.ExcelWriter(f'{string}.xlsx', engine='openpyxl') as writer:
#         for sheet_name, split_df in split_sheets_dict.items():
#             split_df.to_excel(writer, sheet_name=sheet_name, index=False)


for string in strings:
    # 打开原始的模板Excel文件和目标Excel文件
    template_file = '2023年Q1业绩综合分析V4终版-4.11.xlsx'
    destination_file = f'{string}.xlsx'

    # 加载工作簿
    template_wb = load_workbook(template_file)
    destination_wb = load_workbook(destination_file)

    # 获取源工作表和目标工作表
    template_sheets = template_wb.sheetnames
    destination_sheets = destination_wb.sheetnames

    # 遍历每个工作表
    for i, template_sheet_name in enumerate(template_sheets):
        destination_sheet_name = destination_sheets[i]  # 目标工作表和源工作表顺序相同
        template_sheet = template_wb[template_sheet_name]
        destination_sheet = destination_wb[destination_sheet_name]

        # 获取源工作表的列数
        column_count = template_sheet.max_column

        # 复制每一列的样式到目标工作表
        for col in range(1, column_count + 1):
            col_letter = get_column_letter(col)  # 列索引转换为字母表示的列名，如1->A, 2->B
            source_column = template_sheet[col_letter]
            destination_column = destination_sheet[col_letter]

            for source_cell, destination_cell in zip(source_column, destination_column):
                destination_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    underline=source_cell.font.underline,
                    color=source_cell.font.color
                )

                destination_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    fgColor=source_cell.fill.fgColor,
                    bgColor=source_cell.fill.bgColor
                )

                destination_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )

                destination_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrapText=source_cell.alignment.wrapText
                )

                destination_cell.number_format = source_cell.number_format

    # 保存目标工作簿文件
    destination_wb.save(destination_file)
